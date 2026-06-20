"""
AWS Pricing BOQ Generator - Dixit Infotech
Streamlit app: pull line items from an AWS Pricing Calculator estimate
(by link, or by an exported CSV), add manual commercial inputs
(MSP, FW, EDR, one-time deployment, dollar rate, note, T&C),
and download a fully formatted Excel BOQ matching the Dixit Infotech template.
"""

import csv
import io
import math
import re
from datetime import date

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------------------

READ_API_BASE = "https://d3knqfixx3sbls.cloudfront.net"  # public, unauthenticated calculator.aws read endpoint

DEFAULT_TERMS = [
    "* AWS Pricing Calculator provides only an estimate of your AWS fees and doesn't include any taxes that might apply. Your actual fees depend on a variety of factors, including your actual usage of AWS services.",
    "The currency for all pricing, either US Dollars or Indian Rupees, is dependent on the Original Equipment Manufacturer (OEM).",
    "Taxes will be applicable at actuals",
    "The customer is responsible for any tax, duty, or levy increases resulting from changes in government policies.",
    "Proposed commercials is a estimate. However actual billing would be based on the Cloud utilisation month on month basis",
    "Any software, activities, or services not specified within the Bill of Materials (BOM) are considered outside the purview of Dixit Infotech's services. Should the client require such items, separate cost agreements will be necessary.",
    "The client is responsible for the installation, management, and deployment of the application and database.",
    "To mitigate the risk of data loss, backup solutions are strongly advised. In cases where backup services are not incorporated into the commercial terms, the customer is required to subscribe separately. Dixit Infotech shall not be held liable for any data loss incurred if the customer chooses not to subscribe to backup services.",
    "To protect against viruses, worms, malware, ransomware, and network security threats, antivirus software and a firewall are required. If these are not implemented, the client must subscribe to a separate security solution",
    "If additional resources or components are needed, charges will be applied as actually necessary. Additionally, prices are prone to swings in the high or low end depending on the pricing standard used in the market.",
    "Any one-time deployment changes are expressly excluded from the scope of these estimations. As prescribed by the Statement of Work (SOW), one-time setup and migration costs shall be levied as supplementary charges and, if applicable, shall be set forth in the proposal.",
    "The provided commercial proposal is an estimation. Actual billing will be determined by the monthly consumption of cloud resources.",
    "Dixit Infotech will act as a partner to facilitate reporting, billing dashboards, consumption analysis, and health checks for all subscriptions.",
    "Billing commences upon infrastructure provisioning, regardless of the project's Go-Live date.",
    "The project's Statement of Work (SOW) will be provided separately and requires mutual agreement.",
    "Compliance with all applicable OEM norms, policies, and regulations shall be maintained for the Reserved Instances procured",
]

DETAIL_COLUMNS = [
    "Group hierarchy",
    "Region",
    "Description",
    "Service",
    "Upfront",
    "Monthly",
    "Currency",
    "Configuration summary",
]

EMPTY_DETAIL_DF = pd.DataFrame(columns=DETAIL_COLUMNS)

# --------------------------------------------------------------------------------------
# AWS Pricing Calculator link -> data extraction
# --------------------------------------------------------------------------------------


def extract_estimate_id(link_or_id: str) -> str:
    """Pull the estimate id out of a calculator.aws share link, or pass through if already an id."""
    link_or_id = (link_or_id or "").strip()
    m = re.search(r"[?&#]id=([A-Za-z0-9]+)", link_or_id)
    if m:
        return m.group(1)
    return link_or_id


def _num(v, default=0.0):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def fetch_estimate_json(estimate_id: str) -> dict:
    url = f"{READ_API_BASE}/{estimate_id}"
    resp = requests.get(url, timeout=20)
    resp.raise_for_status()
    return resp.json()


def rows_from_json(data: dict) -> list:
    """Best-effort extraction of line items from the calculator.aws estimate JSON.
    This endpoint is public but undocumented, so this parser is defensive and will
    happily return an empty list rather than crash -- use the CSV upload tab as a
    guaranteed-accurate fallback if a future AWS change breaks this."""
    rows = []
    top_name = data.get("name") or data.get("title") or "AWS Estimate"

    def walk(node):
        if not isinstance(node, dict):
            return
        services = node.get("services")
        if isinstance(services, dict):
            for _, svc in services.items():
                if not isinstance(svc, dict):
                    continue
                cost = svc.get("serviceCost") or svc.get("cost") or {}
                monthly = _num(cost.get("monthly", cost.get("monthlyCost")))
                upfront = _num(cost.get("upfront", cost.get("upfrontCost")))
                rows.append(
                    {
                        "Group hierarchy": top_name,
                        "Region": svc.get("regionName") or svc.get("region") or "",
                        "Description": svc.get("description") or "",
                        "Service": svc.get("serviceName") or svc.get("name") or "",
                        "Upfront": upfront,
                        "Monthly": monthly,
                        "Currency": "USD",
                        "Configuration summary": svc.get("configSummary") or svc.get("description") or "",
                    }
                )
        elif isinstance(services, list):
            for svc in services:
                if not isinstance(svc, dict):
                    continue
                cost = svc.get("serviceCost") or svc.get("cost") or {}
                monthly = _num(cost.get("monthly", cost.get("monthlyCost")))
                upfront = _num(cost.get("upfront", cost.get("upfrontCost")))
                rows.append(
                    {
                        "Group hierarchy": top_name,
                        "Region": svc.get("regionName") or svc.get("region") or "",
                        "Description": svc.get("description") or "",
                        "Service": svc.get("serviceName") or svc.get("name") or "",
                        "Upfront": upfront,
                        "Monthly": monthly,
                        "Currency": "USD",
                        "Configuration summary": svc.get("configSummary") or svc.get("description") or "",
                    }
                )
        groups = node.get("groups")
        if isinstance(groups, dict):
            for _, grp in groups.items():
                walk(grp)
        elif isinstance(groups, list):
            for grp in groups:
                walk(grp)

    walk(data)
    return rows


# --------------------------------------------------------------------------------------
# CSV export (the file produced by AWS Pricing Calculator's own Export -> CSV button)
# --------------------------------------------------------------------------------------


def rows_from_csv(file_bytes: bytes) -> list:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = list(csv.reader(io.StringIO(text)))

    header_idx = None
    for i, row in enumerate(reader):
        norm = [c.strip().lower() for c in row]
        if "group hierarchy" in norm and "service" in norm:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            "Could not find the 'Detailed Estimate' table in this CSV "
            "(expected a header row containing 'Group hierarchy' and 'Service'). "
            "Make sure this is the CSV exported from AWS Pricing Calculator's Export button."
        )

    header = [c.strip() for c in reader[header_idx]]
    col_idx = {name.lower(): idx for idx, name in enumerate(header)}

    def get(row, name, default=""):
        idx = col_idx.get(name.lower())
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    rows = []
    for row in reader[header_idx + 1 :]:
        if not any(c.strip() for c in row):
            break
        if len(row) < 4:
            continue
        upfront = _num(get(row, "Upfront", 0))
        monthly = _num(get(row, "Monthly", 0))
        rows.append(
            {
                "Group hierarchy": get(row, "Group hierarchy"),
                "Region": get(row, "Region"),
                "Description": get(row, "Description"),
                "Service": get(row, "Service"),
                "Upfront": upfront,
                "Monthly": monthly,
                "Currency": get(row, "Currency", "USD") or "USD",
                "Configuration summary": get(row, "Configuration summary"),
            }
        )
    return rows


# --------------------------------------------------------------------------------------
# Excel workbook builder
# --------------------------------------------------------------------------------------

FONT_NAME = "Calibri"
TITLE_FILL = PatternFill("solid", fgColor="DAE3F3")   # blue accent5, lighter 80%
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")  # blue accent1, lighter 60%
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")    # gold accent4, lighter 80%
USD_FMT = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
INR_FMT = '"\u20b9"#,##0.00_);[Red]\\("\u20b9"#,##0.00\\)'
THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")


def _border(top, bottom, left, right):
    m = {"thin": THIN, "medium": MEDIUM, None: None}
    return Border(top=m[top], bottom=m[bottom], left=m[left], right=m[right])


def grid_border(ws, min_row, max_row, min_col, max_col, outer="medium", inner="thin"):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            top = outer if r == min_row else inner
            bottom = outer if r == max_row else inner
            left = outer if c == min_col else inner
            right = outer if c == max_col else inner
            cell.border = _border(top, bottom, left, right)


def fill_range(ws, min_row, max_row, min_col, max_col, fill):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).fill = fill


def set_cell(ws, coord, value, bold=False, fill=None, fmt=None, align=None,
             wrap=False, font_color=None, size=11, vertical="center"):
    c = ws[coord]
    c.value = value
    c.font = Font(name=FONT_NAME, size=size, bold=bold, color=font_color)
    if fill is not None:
        c.fill = fill
    if fmt is not None:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical=vertical, wrap_text=wrap)
    return c


def estimate_row_height(text: str, chars_per_line: int = 80, line_height: int = 15, min_height: int = 30) -> float:
    if not text:
        return min_height
    lines = max(1, math.ceil(len(str(text)) / chars_per_line)) + 1  # +1 line buffer to avoid clipped wrap text
    return max(min_height, lines * line_height)


def build_workbook(
    detail_df: pd.DataFrame,
    aws_link: str,
    dollar_rate: float,
    msp_percent: float,
    onetime_total: float,
    include_fw: bool,
    fw_total: float,
    include_edr: bool,
    edr_total: float,
    include_note: bool,
    note_text: str,
    include_terms: bool,
    terms_lines: list,
) -> io.BytesIO:

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.sheet_view.showGridLines = False  # blank sheet area stays white; only explicit cell borders show

    widths = {"A": 5, "B": 23, "C": 17, "D": 17, "E": 10, "F": 10,
              "G": 13, "H": 17, "I": 12, "J": 75, "K": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["F"].hidden = True  # Upfront column hidden by default (all costs typically $0 upfront)

    n_detail = len(detail_df)

    # ---- pre-compute the row layout so forward-referencing formulas are easy ----
    row_title = 2
    row_aws_hdr = 3
    row_aws_data = 4
    row_summary_hdr = 6
    row_aws_total = 7
    row_msp = 8
    row_onetime = 9
    next_row = 10
    row_fw = next_row if include_fw else None
    if include_fw:
        next_row += 1
    row_edr = next_row if include_edr else None
    if include_edr:
        next_row += 1
    row_total_estimate = next_row
    row_detail_title = row_total_estimate + 3  # 2 blank rows before Detailed Estimate
    row_detail_hdr = row_detail_title + 1
    first_detail_row = row_detail_hdr + 1
    last_detail_row = first_detail_row + max(n_detail, 1) - 1

    cursor = last_detail_row + 3  # 2 blank rows before the next section

    row_note_title = None
    if include_note:
        row_note_title = cursor
        cursor = row_note_title + 2 + 2  # note block is 2 rows tall, then 2 blank rows before the next section

    row_terms_title = None
    row_terms_first = None
    row_terms_last = None
    if include_terms and terms_lines:
        row_terms_title = cursor
        row_terms_first = row_terms_title + 1
        row_terms_last = row_terms_first + len(terms_lines) - 1
        cursor = row_terms_last + 1

    # ============================ Estimate summary ============================
    ws.merge_cells(f"B{row_title}:E{row_title}")
    set_cell(ws, f"B{row_title}", "Estimate summary", bold=True, fill=TITLE_FILL, align="left")
    fill_range(ws, row_title, row_title, 2, 5, TITLE_FILL)
    grid_border(ws, row_title, row_title, 2, 5, outer="medium", inner="medium")

    ws.merge_cells(f"H{row_title}:J{row_title}")
    set_cell(ws, f"H{row_title}", "AWS BOQ Link", bold=True, fill=TITLE_FILL, align="left")
    fill_range(ws, row_title, row_title, 8, 10, TITLE_FILL)
    grid_border(ws, row_title, row_title, 8, 10, outer="medium", inner="medium")

    for col, h in zip("BCDE", ["Upfront cost", "Monthly cost", "Total 12 months cost", "Currency"]):
        set_cell(ws, f"{col}{row_aws_hdr}", h, bold=True, fill=HEADER_FILL, align="center", wrap=True)
    grid_border(ws, row_aws_hdr, row_aws_hdr, 2, 5, outer="medium", inner="thin")

    ws.merge_cells(f"H{row_aws_hdr}:J{row_aws_hdr}")
    set_cell(ws, f"H{row_aws_hdr}", aws_link or "", align="left", font_color="FF0000FF")
    fill_range(ws, row_aws_hdr, row_aws_hdr, 8, 10, PatternFill(fill_type=None))
    grid_border(ws, row_aws_hdr, row_aws_hdr, 8, 10, outer="thin", inner="thin")

    set_cell(ws, f"B{row_aws_data}", f"=SUM(F{first_detail_row}:F{last_detail_row})", align="center", fmt=USD_FMT)
    set_cell(ws, f"C{row_aws_data}", f"=SUM(G{first_detail_row}:G{last_detail_row})", align="center", fmt=USD_FMT)
    set_cell(ws, f"D{row_aws_data}", f"=B{row_aws_data}+C{row_aws_data}*12", align="center", fmt=USD_FMT)
    set_cell(ws, f"E{row_aws_data}", "USD", align="center")
    grid_border(ws, row_aws_data, row_aws_data, 2, 5, outer="medium", inner="thin")

    # ============================ Summary (INR) ============================
    for col, h in zip("BCDE", ["Summary", "Monthly cost", "Total 12 months cost", "Currency"]):
        set_cell(ws, f"{col}{row_summary_hdr}", h, bold=True, fill=HEADER_FILL, align="center", wrap=True)
    grid_border(ws, row_summary_hdr, row_summary_hdr, 2, 5, outer="medium", inner="thin")

    set_cell(ws, f"H{row_summary_hdr}", "USD_INR", align="left")
    set_cell(ws, f"I{row_summary_hdr}", dollar_rate, align="center")
    set_cell(ws, f"J{row_summary_hdr}", "Estimated Dollar Rate, will be charged as per OEM Invoice", align="left")

    dollar_rate_cell = f"$I${row_summary_hdr}"
    wb.defined_names["USD_INR"] = DefinedName("USD_INR", attr_text=f"Sheet1!{dollar_rate_cell}")

    set_cell(ws, f"B{row_aws_total}", "AWS Total Estimate", align="left")
    set_cell(ws, f"C{row_aws_total}", f"=C{row_aws_data}*USD_INR", fmt=INR_FMT, align="center")
    set_cell(ws, f"D{row_aws_total}", f"=D{row_aws_data}*USD_INR", fmt=INR_FMT, align="center")
    set_cell(ws, f"E{row_aws_total}", "INR", align="center")

    set_cell(ws, f"B{row_msp}", "MSP", align="left")
    set_cell(ws, f"C{row_msp}", f"=C{row_aws_total}*I{row_msp}/100", fmt=INR_FMT, align="center")
    set_cell(ws, f"D{row_msp}", f"=C{row_msp}*12", fmt=INR_FMT, align="center")
    set_cell(ws, f"E{row_msp}", "INR", align="center")
    set_cell(ws, f"H{row_msp}", "MSP %", align="left")
    set_cell(ws, f"I{row_msp}", msp_percent, fmt='0.00"%"', align="center", font_color="FF0000FF")
    set_cell(ws, f"J{row_msp}", "Percentage of AWS Total Estimate (Monthly, INR) charged as MSP fee", align="left")

    set_cell(ws, f"B{row_onetime}", "One time Deployemnt & Migration", align="left")
    set_cell(ws, f"C{row_onetime}", "-", fmt=INR_FMT, align="center")
    set_cell(ws, f"D{row_onetime}", onetime_total, fmt=INR_FMT, align="center", font_color="FF0000FF")
    set_cell(ws, f"E{row_onetime}", "INR", align="center")

    if include_fw:
        set_cell(ws, f"B{row_fw}", "FW License (Annual)", align="left")
        set_cell(ws, f"C{row_fw}", "-", fmt=INR_FMT, align="center")
        set_cell(ws, f"D{row_fw}", fw_total, fmt=INR_FMT, align="center", font_color="FF0000FF")
        set_cell(ws, f"E{row_fw}", "INR", align="center")

    if include_edr:
        set_cell(ws, f"B{row_edr}", "EDR License (Annual)", align="left")
        set_cell(ws, f"C{row_edr}", "-", fmt=INR_FMT, align="center")
        set_cell(ws, f"D{row_edr}", edr_total, fmt=INR_FMT, align="center", font_color="FF0000FF")
        set_cell(ws, f"E{row_edr}", "INR", align="center")

    set_cell(ws, f"B{row_total_estimate}", "Total Estimate", bold=True, fill=HEADER_FILL, align="left")
    set_cell(ws, f"C{row_total_estimate}", f"=SUM(C{row_aws_total}:C{row_total_estimate - 1})",
             bold=True, fill=HEADER_FILL, fmt=INR_FMT, align="center")
    set_cell(ws, f"D{row_total_estimate}", f"=SUM(D{row_aws_total}:D{row_total_estimate - 1})",
             bold=True, fill=HEADER_FILL, fmt=INR_FMT, align="center")
    set_cell(ws, f"E{row_total_estimate}", "INR", bold=True, fill=HEADER_FILL, align="center")

    grid_border(ws, row_aws_total, row_total_estimate, 2, 5, outer="medium", inner="thin")

    # ============================ Detailed estimate ============================
    ws.merge_cells(f"B{row_detail_title}:J{row_detail_title}")
    set_cell(ws, f"B{row_detail_title}", "Detailed Estimate", bold=True, fill=TITLE_FILL, align="left")
    fill_range(ws, row_detail_title, row_detail_title, 2, 10, TITLE_FILL)
    grid_border(ws, row_detail_title, row_detail_title, 2, 10, outer="medium", inner="medium")

    detail_headers = ["Group hierarchy", "Region", "Description", "Service",
                       "Upfront", "Monthly", "First 12 months total", "Currency", "Configuration summary"]
    for col, h in zip("BCDEFGHIJ", detail_headers):
        set_cell(ws, f"{col}{row_detail_hdr}", h, bold=True, fill=HEADER_FILL, align="center", wrap=True)
    grid_border(ws, row_detail_hdr, row_detail_hdr, 2, 10, outer="medium", inner="thin")

    if n_detail == 0:
        r = first_detail_row
        for col in "BCDEFGHIJ":
            set_cell(ws, f"{col}{r}", "" if col != "I" else "USD", align="center", wrap=True)
        ws.row_dimensions[r].height = 30
    else:
        for i, (_, row) in enumerate(detail_df.iterrows()):
            r = first_detail_row + i
            set_cell(ws, f"B{r}", row.get("Group hierarchy", ""), wrap=True)
            set_cell(ws, f"C{r}", row.get("Region", ""), wrap=True)
            set_cell(ws, f"D{r}", row.get("Description", ""), wrap=True)
            set_cell(ws, f"E{r}", row.get("Service", ""), wrap=True)
            set_cell(ws, f"F{r}", _num(row.get("Upfront", 0)), align="center")
            set_cell(ws, f"G{r}", _num(row.get("Monthly", 0)), fmt=USD_FMT, align="center")
            set_cell(ws, f"H{r}", f"=F{r}+G{r}*12", fmt=USD_FMT, align="center")
            set_cell(ws, f"I{r}", row.get("Currency", "USD") or "USD", align="center")
            set_cell(ws, f"J{r}", row.get("Configuration summary", ""), wrap=True, align="left")
            cfg = str(row.get("Configuration summary", "") or "")
            ws.row_dimensions[r].height = estimate_row_height(cfg)

    grid_border(ws, first_detail_row, last_detail_row, 2, 10, outer="medium", inner="thin")

    # ============================ Note ============================
    if include_note:
        ws.merge_cells(f"B{row_note_title}:B{row_note_title + 1}")
        set_cell(ws, f"B{row_note_title}", "Note (Optional)", bold=True, fill=NOTE_FILL, align="left", wrap=True)
        fill_range(ws, row_note_title, row_note_title + 1, 2, 2, NOTE_FILL)
        ws.merge_cells(f"C{row_note_title}:G{row_note_title + 1}")
        set_cell(ws, f"C{row_note_title}", note_text or "", align="left", wrap=True)
        grid_border(ws, row_note_title, row_note_title + 1, 2, 7, outer="medium", inner="thin")
        ws.row_dimensions[row_note_title].height = max(20, estimate_row_height(note_text, chars_per_line=60) / 2)
        ws.row_dimensions[row_note_title + 1].height = max(20, estimate_row_height(note_text, chars_per_line=60) / 2)

    # ============================ Terms and Conditions ============================
    if include_terms and terms_lines:
        ws.merge_cells(f"B{row_terms_title}:O{row_terms_title}")
        set_cell(ws, f"B{row_terms_title}", "Terms and Conditions", bold=True, fill=NOTE_FILL, align="left")
        fill_range(ws, row_terms_title, row_terms_title, 2, 15, NOTE_FILL)
        grid_border(ws, row_terms_title, row_terms_title, 2, 15, outer="medium", inner="medium")

        for i, line in enumerate(terms_lines):
            r = row_terms_first + i
            ws.merge_cells(f"B{r}:O{r}")
            set_cell(ws, f"B{r}", line, align="left", wrap=True)
            fill_range(ws, r, r, 2, 15, PatternFill(fill_type=None))
            ws.row_dimensions[r].height = estimate_row_height(line, chars_per_line=160)
        grid_border(ws, row_terms_first, row_terms_last, 2, 15, outer="medium", inner="thin")

    ws.row_dimensions[1].height = 14.5
    ws.row_dimensions[row_title].height = 18
    ws.row_dimensions[row_aws_hdr].height = 28
    ws.row_dimensions[row_aws_data].height = 22
    ws.row_dimensions[row_summary_hdr].height = 28
    ws.row_dimensions[row_detail_hdr].height = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------------------
# Streamlit UI
# --------------------------------------------------------------------------------------

st.set_page_config(page_title="AWS Pricing BOQ Generator - Dixit Infotech", layout="wide")

if "detail_df" not in st.session_state:
    st.session_state.detail_df = EMPTY_DETAIL_DF.copy()
if "aws_link" not in st.session_state:
    st.session_state.aws_link = ""
if "form_version" not in st.session_state:
    st.session_state.form_version = 0

FV = st.session_state.form_version  # suffix for widget keys, bumped on reset so every field reverts to default


def reset_app():
    st.session_state.detail_df = EMPTY_DETAIL_DF.copy()
    st.session_state.aws_link = ""
    st.session_state.form_version += 1


title_col, reset_col = st.columns([8, 1])
with title_col:
    st.title("AWS Pricing BOQ Generator")
    st.caption("Dixit Infotech \u2014 build a customer-ready AWS pricing BOQ from an AWS Pricing Calculator estimate")
with reset_col:
    st.write("")
    if st.button("\U0001F504 New", help="Clear everything and start a new proposal", use_container_width=True):
        reset_app()
        st.rerun()

st.header("1. AWS Pricing Calculator estimate")
tab_link, tab_csv = st.tabs(["\U0001F517 Fetch from link", "\U0001F4E4 Upload exported CSV"])

with tab_link:
    link = st.text_input(
        "AWS Pricing Calculator share link",
        value=st.session_state.aws_link,
        placeholder="https://calculator.aws/#/estimate?id=...",
        key=f"link_input_{FV}",
    )
    if st.button("Fetch estimate data", type="primary", key=f"fetch_btn_{FV}"):
        st.session_state.aws_link = link
        estimate_id = extract_estimate_id(link)
        if not estimate_id:
            st.error("Couldn't find an estimate id in that link. Paste the full calculator.aws share link.")
        else:
            try:
                with st.spinner("Fetching estimate from AWS\u2026"):
                    data = fetch_estimate_json(estimate_id)
                    rows = rows_from_json(data)
                if rows:
                    st.session_state.detail_df = pd.DataFrame(rows)[DETAIL_COLUMNS]
                    st.success(f"Pulled {len(rows)} line item(s) from the estimate.")
                else:
                    st.warning(
                        "Connected to AWS but couldn't find any line items in the response. "
                        "This read endpoint is unofficial and AWS can change its shape without notice \u2014 "
                        "use the 'Upload exported CSV' tab instead (Export \u2192 CSV on the calculator.aws page) "
                        "for a guaranteed-accurate result."
                    )
                with st.expander("Raw API response (debug)"):
                    st.json(data)
            except Exception as e:
                st.error(
                    f"Couldn't fetch the estimate automatically ({e}). "
                    "Use the 'Upload exported CSV' tab instead: open the link in your browser, "
                    "click Export \u2192 CSV on the My Estimate page, and upload that file here."
                )

with tab_csv:
    st.write("On the AWS Pricing Calculator page (My estimate), use **Export \u2192 CSV**, then upload that file here.")
    uploaded = st.file_uploader("AWS Pricing Calculator CSV export", type=["csv"], key=f"uploader_{FV}")
    if uploaded is not None:
        try:
            rows = rows_from_csv(uploaded.getvalue())
            st.session_state.detail_df = pd.DataFrame(rows)[DETAIL_COLUMNS]
            st.success(f"Loaded {len(rows)} line item(s) from the CSV.")
        except Exception as e:
            st.error(str(e))

st.subheader("Detailed Estimate line items")
st.caption("Review, edit, add, or remove rows as needed \u2014 this becomes the Detailed Estimate table.")
edited = st.data_editor(
    st.session_state.detail_df,
    num_rows="dynamic",
    use_container_width=True,
    key=f"detail_editor_{FV}",
    column_config={
        "Upfront": st.column_config.NumberColumn(format="%.2f"),
        "Monthly": st.column_config.NumberColumn(format="%.2f"),
    },
)
st.session_state.detail_df = edited

if not edited.empty:
    total_monthly = pd.to_numeric(edited["Monthly"], errors="coerce").fillna(0).sum()
    total_upfront = pd.to_numeric(edited["Upfront"], errors="coerce").fillna(0).sum()
    c1, c2, c3 = st.columns(3)
    c1.metric("Upfront (USD)", f"${total_upfront:,.2f}")
    c2.metric("Monthly (USD)", f"${total_monthly:,.2f}")
    c3.metric("12-month total (USD)", f"${total_upfront + total_monthly * 12:,.2f}")

st.divider()

st.header("2. Commercial inputs (manual)")
col1, col2 = st.columns(2)
with col1:
    dollar_rate = st.number_input("Dollar rate (USD \u2192 INR)", min_value=0.0, value=93.0, step=0.1,
                                   format="%.2f", key=f"dollar_rate_{FV}")
    msp_percent = st.number_input("MSP \u2014 % of AWS Total Estimate (Monthly)", min_value=0.0, value=10.0,
                                   step=0.5, format="%.2f", key=f"msp_percent_{FV}")
    onetime_total = st.number_input("One time Deployment & Migration \u2014 Total (INR)", min_value=0.0, value=0.0,
                                     step=1000.0, format="%.2f", key=f"onetime_{FV}")
with col2:
    include_fw = st.checkbox("Include FW License (Annual)", value=True, key=f"include_fw_{FV}")
    fw_total = st.number_input("FW License \u2014 Total (INR)", min_value=0.0, value=0.0, step=1000.0,
                                format="%.2f", disabled=not include_fw, key=f"fw_total_{FV}")
    include_edr = st.checkbox("Include EDR License (Annual)", value=True, key=f"include_edr_{FV}")
    edr_total = st.number_input("EDR License \u2014 Total (INR)", min_value=0.0, value=0.0, step=1000.0,
                                 format="%.2f", disabled=not include_edr, key=f"edr_total_{FV}")

st.divider()

st.header("3. Note (optional)")
include_note = st.checkbox("Include Note section", value=True, key=f"include_note_{FV}")
note_text = st.text_area("Note text", value="", disabled=not include_note, height=80, key=f"note_text_{FV}")

st.divider()

st.header("4. Terms and Conditions (optional)")
include_terms = st.checkbox("Include Terms and Conditions section", value=True, key=f"include_terms_{FV}")
terms_text = st.text_area(
    "One term per line \u2014 edit, remove, or add lines as needed",
    value="\n".join(DEFAULT_TERMS),
    disabled=not include_terms,
    height=260,
    key=f"terms_text_{FV}",
)

st.divider()

st.header("5. Generate")
project_name = st.text_input("Project / file name", value="AWS-Pricing-Estimate", key=f"project_name_{FV}")

if st.button("Generate Excel BOQ", type="primary", key=f"generate_btn_{FV}"):
    terms_lines = [l.strip() for l in terms_text.split("\n") if l.strip()] if include_terms else []
    buf = build_workbook(
        detail_df=st.session_state.detail_df,
        aws_link=st.session_state.aws_link,
        dollar_rate=dollar_rate,
        msp_percent=msp_percent,
        onetime_total=onetime_total,
        include_fw=include_fw,
        fw_total=fw_total,
        include_edr=include_edr,
        edr_total=edr_total,
        include_note=include_note,
        note_text=note_text,
        include_terms=include_terms,
        terms_lines=terms_lines,
    )
    fname = f"{project_name.strip().replace(' ', '-') or 'AWS-Pricing-Estimate'}_{date.today().isoformat()}.xlsx"
    st.download_button(
        "\u2b07\ufe0f Download Excel BOQ",
        data=buf,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        key=f"download_btn_{FV}",
    )
    st.success("Workbook generated \u2014 click the button above to download.")